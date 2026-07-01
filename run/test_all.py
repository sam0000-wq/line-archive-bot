# -*- coding: utf-8 -*-
"""
test_all.py - Automated test suite for LINE Archive Bot.
Usage:
    python test_all.py              # run all tests
    python test_all.py --verbose    # detailed output
"""

import argparse
import logging
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))
logging.basicConfig(level=logging.CRITICAL)


class TestResult:
    def __init__(self, name: str, description: str):
        self.name = name
        self.description = description
        self.passed: Optional[bool] = None
        self.message: str = ""
        self.duration_ms: float = 0.0


def test_decorator(name: str, description: str = ""):
    """Decorator factory for test methods."""
    def decorator(func: Callable) -> Callable:
        def wrapper(self) -> TestResult:
            result = TestResult(name, description or func.__doc__ or name)
            t0 = time.perf_counter()
            try:
                func(self, result)
                result.passed = True
                if not result.message:
                    result.message = "OK"
            except AssertionError as e:
                result.passed = False
                result.message = str(e) or "Assertion failed"
            except Exception as e:
                result.passed = False
                result.message = f"{type(e).__name__}: {e}\n{traceback.format_exc()}"
            result.duration_ms = (time.perf_counter() - t0) * 1000
            return result
        wrapper._test_meta = {"name": name, "description": description}
        wrapper._is_test = True
        return wrapper
    return decorator


class TestRunner:
    def __init__(self):
        self.results: list[TestResult] = []

    def run(self) -> list[TestResult]:
        for attr_name in dir(self):
            attr = getattr(self, attr_name)
            if getattr(attr, "_is_test", False):
                result = attr()
                self.results.append(result)
        return self.results


class ConfigTests(TestRunner):
    @test_decorator("Load config module")
    def test_load_config(self, result: TestResult) -> None:
        from config import Config
        assert hasattr(Config, "LINE_CHANNEL_SECRET"), "Missing LINE_CHANNEL_SECRET"
        assert hasattr(Config, "GMAIL_USER"), "Missing GMAIL_USER"
        assert Config.GMAIL_USER == "itsamliu2025@gmail.com", f"Expected itsamliu2025@gmail.com, got {Config.GMAIL_USER}"
        assert Config.TZ == "Asia/Taipei", f"Expected Asia/Taipei, got {Config.TZ}"

    @test_decorator("Config directory creation")
    def test_ensure_dirs(self, result: TestResult) -> None:
        from config import Config
        import tempfile
        import shutil
        original = Config.ARCHIVE_DIR
        Config.ARCHIVE_DIR = Path(tempfile.mkdtemp())
        try:
            Config.ensure_dirs()
            assert Config.ARCHIVE_DIR.exists(), "ARCHIVE_DIR not created"
            assert Config.LOG_DIR.exists(), "LOG_DIR not created"
        finally:
            shutil.rmtree(str(Config.ARCHIVE_DIR.parent), ignore_errors=True)
            Config.ARCHIVE_DIR = original

    @test_decorator("Config validation - missing vars detected")
    def test_validate_missing(self, result: TestResult) -> None:
        from config import Config
        orig_secret = Config.LINE_CHANNEL_SECRET
        orig_token = Config.LINE_CHANNEL_ACCESS_TOKEN
        orig_password = Config.GMAIL_APP_PASSWORD
        Config.LINE_CHANNEL_SECRET = ""
        Config.LINE_CHANNEL_ACCESS_TOKEN = ""
        Config.GMAIL_APP_PASSWORD = ""
        missing = Config.validate()
        assert "LINE_CHANNEL_SECRET" in missing
        assert "LINE_CHANNEL_ACCESS_TOKEN" in missing
        assert "GMAIL_APP_PASSWORD" in missing
        Config.LINE_CHANNEL_SECRET = orig_secret
        Config.LINE_CHANNEL_ACCESS_TOKEN = orig_token
        Config.GMAIL_APP_PASSWORD = orig_password


class ArchiveTests(TestRunner):
    @test_decorator("Message classification - critical")
    def test_classify_critical(self, result: TestResult) -> None:
        from archiver import _get_sheet_name
        assert _get_sheet_name("critical: server down") == "critical"
        assert _get_sheet_name("CRITICAL: urgent") == "critical"
        assert _get_sheet_name("  Critical: test") == "critical"

    @test_decorator("Message classification - warning")
    def test_classify_warning(self, result: TestResult) -> None:
        from archiver import _get_sheet_name
        assert _get_sheet_name("warning: disk full") == "warning"
        assert _get_sheet_name("WARNING: alert") == "warning"
        assert _get_sheet_name("  Warning: test") == "warning"

    @test_decorator("Message classification - others")
    def test_classify_others(self, result: TestResult) -> None:
        from archiver import _get_sheet_name
        assert _get_sheet_name("hello world") == "others"
        assert _get_sheet_name("") == "others"
        assert _get_sheet_name("critical") == "others"
        assert _get_sheet_name("warning") == "others"

    @test_decorator("XLSX creation and archiving")
    def test_archive_xlsx(self, result: TestResult) -> None:
        from archiver import archive_message
        from config import Config
        import tempfile
        import shutil
        from openpyxl import load_workbook
        original_dir = Config.ARCHIVE_DIR
        Config.ARCHIVE_DIR = Path(tempfile.mkdtemp())
        Config.ensure_dirs()
        try:
            sheet = archive_message(timestamp="2025-07-01 12:00:00", sender_name="TestUser", sender_user_id="Utest123", message="critical: system error occurred")
            assert sheet == "critical", f"Expected critical, got {sheet}"
            files = list(Config.ARCHIVE_DIR.glob("*.xlsx"))
            assert len(files) == 1, f"Expected 1 xlsx, found {len(files)}"
            wb = load_workbook(str(files[0]))
            assert "critical" in wb.sheetnames
            assert "warning" in wb.sheetnames
            assert "others" in wb.sheetnames
            ws = wb["critical"]
            assert ws.max_row == 2, f"Expected 2 rows, got {ws.max_row}"
            assert ws.cell(2, 4).value == "critical: system error occurred"
            archive_message(timestamp="2025-07-01 13:00:00", sender_name="User2", sender_user_id="Utest456", message="warning: check this")
            wb = load_workbook(str(files[0]))
            ws_warn = wb["warning"]
            assert ws_warn.max_row == 2
            result.message = f"Created {files[0].name} with 2 messages across 2 sheets"
        finally:
            shutil.rmtree(str(Config.ARCHIVE_DIR), ignore_errors=True)
            Config.ARCHIVE_DIR = original_dir


class MailerTests(TestRunner):
    @test_decorator("Email message building with attachment")
    def test_build_message_with_attachment(self, result: TestResult) -> None:
        from mailer import _build_message
        from config import Config
        import tempfile
        from email.mime.multipart import MIMEMultipart
        tmp = Path(tempfile.mktemp(suffix=".xlsx"))
        tmp.write_text("fake content", encoding="utf-8")
        msg = _build_message(attachment_path=tmp, date_str="20250701")
        assert isinstance(msg, MIMEMultipart)
        assert msg["From"] == Config.GMAIL_USER
        assert "20250701" in msg["Subject"]
        tmp.unlink(missing_ok=True)

    @test_decorator("Email message building without attachment")
    def test_build_message_no_attachment(self, result: TestResult) -> None:
        from mailer import _build_message
        from email.mime.multipart import MIMEMultipart
        msg = _build_message(attachment_path=None, date_str="20250701")
        assert isinstance(msg, MIMEMultipart)
        body = ""
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                body = part.get_payload(decode=True).decode("utf-8")
                break
        assert "No messages were recorded" in body, f"Body text: {body[:200]}"


class FlaskTests(TestRunner):
    @test_decorator("Flask app health endpoint")
    def test_health_endpoint(self, result: TestResult) -> None:
        from app import app
        client = app.test_client()
        resp = client.get("/health")
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        data = resp.get_json()
        assert data.get("status") == "ok"

    @test_decorator("Webhook callback - invalid signature")
    def test_invalid_signature(self, result: TestResult) -> None:
        from app import app
        client = app.test_client()
        resp = client.post("/callback", data='{"events":[]}', content_type="application/json", headers={"X-Line-Signature": "bad"})
        assert resp.status_code == 400

    @test_decorator("Manual report endpoint")
    def test_send_report_endpoint(self, result: TestResult) -> None:
        from app import app
        client = app.test_client()
        resp = client.post("/send-report")
        assert resp.status_code in (200, 500)
        assert resp.get_json() is not None


class IntegrationTests(TestRunner):
    @test_decorator("Config -> Archive -> File: full flow")
    def test_full_archive_flow(self, result: TestResult) -> None:
        from config import Config
        from archiver import archive_message, _daily_path
        import tempfile
        import shutil
        from openpyxl import load_workbook
        original_dir = Config.ARCHIVE_DIR
        Config.ARCHIVE_DIR = Path(tempfile.mkdtemp())
        Config.ensure_dirs()
        try:
            messages = [
                ("2025-07-01 10:00:00", "Alice", "Ualice", "critical: database connection lost"),
                ("2025-07-01 10:01:00", "Bob", "Ubob", "warning: high memory usage"),
                ("2025-07-01 10:02:00", "Charlie", "Ucharlie", "hello everyone"),
            ]
            for ts, name, uid, text in messages:
                sheet = archive_message(ts, name, uid, text)
                assert sheet in ("critical", "warning", "others")
            path = _daily_path()
            assert path.exists()
            wb = load_workbook(str(path))
            assert wb["critical"].max_row == 2
            assert wb["warning"].max_row == 2
            assert wb["others"].max_row == 2
            result.message = f"3 messages archived to {path.name}"
        finally:
            shutil.rmtree(str(Config.ARCHIVE_DIR), ignore_errors=True)
            Config.ARCHIVE_DIR = original_dir


def generate_html_report(results: list[TestResult], output_path: Path) -> str:
    total = len(results)
    passed = sum(1 for r in results if r.passed)
    failed = total - passed
    duration = sum(r.duration_ms for r in results)
    rows = ""
    for r in results:
        status_icon = "&#10004;" if r.passed else "&#10008;"
        status_color = "green" if r.passed else "red"
        cls = "pass" if r.passed else "fail"
        rows += f"""<tr class="{cls}"><td>{r.name}</td><td>{r.description}</td><td style="color:{status_color};font-weight:bold;">{status_icon} {'PASS' if r.passed else 'FAIL'}</td><td>{r.duration_ms:.1f}ms</td><td>{r.message}</td></tr>"""
    overall = "PASS" if passed == total else "FAIL"
    overall_color = "green" if passed == total else "red"
    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8"><title>Test Report - LINE Archive Bot</title>
<style>
* {{ box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; max-width: 1200px; margin: 30px auto; padding: 0 20px; background: #fafafa; }}
h1 {{ color: #333; border-bottom: 2px solid #ddd; }}
.summary {{ display: flex; gap: 20px; margin: 20px 0; flex-wrap: wrap; }}
.summary-card {{ flex: 1; min-width: 120px; padding: 16px; border-radius: 8px; text-align: center; font-size: 1.2em; color: #fff; }}
.card-pass {{ background: #4caf50; }}
.card-fail {{ background: #f44336; }}
.card-total {{ background: #2196f3; }}
.card-time {{ background: #ff9800; }}
table {{ width: 100%; border-collapse: collapse; background: #fff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
th, td {{ text-align: left; padding: 10px 14px; border-bottom: 1px solid #eee; }}
th {{ background: #37474f; color: #fff; }}
.pass td {{ border-left: 4px solid #4caf50; }}
.fail td {{ border-left: 4px solid #f44336; }}
</style>
</head>
<body>
<h1>Test Report - LINE Archive Bot</h1>
<div class="summary">
<div class="summary-card card-pass">&#10004; Passed<br><strong>{passed}/{total}</strong></div>
<div class="summary-card card-fail">&#10008; Failed<br><strong>{failed}</strong></div>
<div class="summary-card card-total">Tests<br><strong>{total}</strong></div>
<div class="summary-card card-time">Duration<br><strong>{duration:.0f}ms</strong></div>
</div>
<div style="text-align:center;margin:16px 0;padding:12px;border-radius:8px;font-size:1.2em;font-weight:bold;color:#fff;background:{overall_color}">OVERALL: {overall}</div>
<table><tr><th>Test Name</th><th>Description</th><th>Status</th><th>Time</th><th>Details</th></tr>{rows}</table>
<p style="color:#888;font-size:0.85em;">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {sys.platform} | Python {sys.version.split()[0]}</p>
</body></html>"""
    output_path.write_text(html, encoding="utf-8")
    return str(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run all tests for LINE Archive Bot")
    parser.add_argument("--verbose", "-v", action="store_true", help="Print detailed output")
    args = parser.parse_args()
    print("LINE Archive Bot - Automated Test Suite")
    print("=" * 56)
    all_results: list[TestResult] = []
    for test_class in [ConfigTests(), ArchiveTests(), MailerTests(), FlaskTests(), IntegrationTests()]:
        class_name = type(test_class).__name__
        print(f"  [{class_name}]")
        for r in test_class.run():
            status = "PASS" if r.passed else "FAIL"
            print(f"    {'+' if r.passed else 'X'} {r.name}: {status} ({r.duration_ms:.0f}ms)")
            if args.verbose or not r.passed:
                print(f"       -> {r.message[:150]}")
        all_results.extend(test_class.results)
    report_path = BASE_DIR / "test_report.html"
    generate_html_report(all_results, report_path)
    total = len(all_results)
    passed = sum(1 for r in all_results if r.passed)
    failed = total - passed
    print("=" * 56)
    print(f"  Results: {passed}/{total} passed, {failed} failed")
    print(f"  HTML Report: {report_path}")
    print("=" * 56)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())


