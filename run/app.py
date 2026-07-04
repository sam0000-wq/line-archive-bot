# -*- coding: utf-8 -*-
"""Flask webhook server for LINE message archiving."""

import json
import logging
import os
import sys
from datetime import datetime
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, jsonify, request, abort, send_file
from linebot.v3 import SignatureValidator
from linebot.v3.webhook import WebhookHandler
from linebot.v3.messaging import MessagingApi, ApiClient, Configuration, TextMessage, PushMessageRequest
from linebot.v3.exceptions import InvalidSignatureError
from linebot.v3.webhooks import MessageEvent, TextMessageContent, JoinEvent
from archiver import archive_message, get_sheet_rows, SHEET_CRITICAL, SHEET_WARNING, SHEET_OTHERS, _get_sheet_name, get_user_display_name, clear_archive
from config import Config
from mailer import send_report
from sync_repo import push_xlsx, pull_xlsx
from llm_analyzer import analyze_messages

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

configuration = Configuration(access_token=Config.LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(Config.LINE_CHANNEL_SECRET)
scheduler = BackgroundScheduler()
TAIPEI_TZ = pytz.timezone("Asia/Taipei")

detected_group_ids: list[str] = []
last_messages: list[dict] = []

monitor = {
    "start_time": datetime.now(TAIPEI_TZ).isoformat(),
    "total_messages_archived": 0,
    "last_message_time": None,
    "errors": [],
    "request_count": 0,
    "last_request_time": None,
    "critical_count": 0,
    "warning_count": 0,
    "last_trigger_time": None,
}


def send_line_report(today: str) -> None:
    if not Config.TARGET_GROUP_ID:
        logger.warning("TARGET_GROUP_ID not set, skipping LINE group notification")
        return
    url = f"{Config.BASE_URL}/files/line_archive_{today}.xlsx"
    msg = f"📋 {today} 群組訊息歸檔報告已產生\n下載: {url}"
    try:
        with ApiClient(configuration) as api_client:
            messaging_api = MessagingApi(api_client)
            messaging_api.push_message(
                PushMessageRequest(
                    to=Config.TARGET_GROUP_ID,
                    messages=[TextMessage(text=msg)],
                )
            )
        logger.info("LINE group notification sent for %s", today)
    except Exception:
        logger.exception("Failed to send LINE group notification for %s", today)


def trigger_instant_report(today: str, trigger_type: str, count: int) -> bool:
    """觸發 GitHub Actions 即時報告 + 發送 LINE 即時通知"""
    # 1. 發送 LINE 即時通知
    if Config.TARGET_GROUP_ID:
        msg = (
            f"🚨 即時觸發報告 ({trigger_type} 達 {count} 筆)\n"
            f"日期: {today}\n"
            f"已觸發 GitHub Actions 寄送 Email (含 xlsx 附件)\n"
            f"下載: {Config.BASE_URL}/files/line_archive_{today}.xlsx"
        )
        try:
            with ApiClient(configuration) as api_client:
                messaging_api = MessagingApi(api_client)
                messaging_api.push_message(
                    PushMessageRequest(
                        to=Config.TARGET_GROUP_ID,
                        messages=[TextMessage(text=msg)],
                    )
                )
            logger.info("Instant LINE notification sent for %s (%s=%d)", today, trigger_type, count)
        except Exception:
            logger.exception("Failed to send instant LINE notification")

    # 2. 觸發 GitHub Actions workflow_dispatch
    if Config.GITHUB_TOKEN:
        import requests
        url = f"https://api.github.com/repos/sam0000-wq/line-archive-bot/actions/workflows/instant-report.yml/dispatches"
        headers = {
            "Authorization": f"Bearer {Config.GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        }
        data = {"ref": "main"}
        try:
            push_xlsx(Config.ARCHIVE_DIR / f"line_archive_{today}.xlsx", today, force=True)
            resp = requests.post(url, headers=headers, json=data, timeout=10)
            if resp.status_code == 204:
                logger.info("GitHub Actions instant-report triggered successfully")
                clear_archive(today)
                return True
            else:
                logger.error("GitHub Actions trigger failed: %d %s", resp.status_code, resp.text)
        except Exception:
            logger.exception("Failed to trigger GitHub Actions")
    return False


def scheduled_report_job() -> None:
    today = datetime.now(TAIPEI_TZ).strftime("%Y%m%d")
    logger.info("Scheduled report triggered for %s", today)
    push_xlsx(Config.ARCHIVE_DIR / f"line_archive_{today}.xlsx", today, force=True)
    success = send_report(today)
    if success:
        logger.info("Scheduled report sent OK for %s", today)
    else:
        logger.error("Scheduled report FAILED for %s", today)
    send_line_report(today)
    clear_archive(today)
    logger.info("Archive cleared for %s after daily report", today)


@app.route("/files/<path:filename>", methods=["GET"])
def serve_file(filename: str):
    from pathlib import Path as _Path
    safe_name = _Path(filename).name
    local_path = Config.ARCHIVE_DIR / safe_name
    
    if not local_path.exists():
        date_str = safe_name.replace("line_archive_", "").replace(".xlsx", "")
        if pull_xlsx(date_str, local_path):
            logger.info("Pulled %s from GitHub for download", safe_name)
        else:
            abort(404, "File not found locally or on GitHub")
    
    return send_file(str(local_path), as_attachment=True)


@app.before_request
def track_request():
    monitor["request_count"] += 1
    monitor["last_request_time"] = datetime.now(TAIPEI_TZ).isoformat()


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "time": datetime.now(TAIPEI_TZ).isoformat()})


@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)
    logger.debug("Webhook received: %s", body[:200])
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        logger.warning("Invalid signature. Ensure webhook URL is set correctly in LINE Developers Console. rejected")
        abort(400, "Invalid signature. Ensure webhook URL is set correctly in LINE Developers Console.")
    return jsonify({"status": "ok"})


@app.route("/group-id", methods=["GET"])
def get_group_id():
    return jsonify({
        "detected_group_ids": list(dict.fromkeys(detected_group_ids)),
        "last_messages": last_messages[-10:],
    })

@app.route("/monitor", methods=["GET"])
def get_monitor():
    uptime = (datetime.now(TAIPEI_TZ) - datetime.fromisoformat(monitor["start_time"])).total_seconds()
    return jsonify({
        "status": "ok",
        "start_time": monitor["start_time"],
        "uptime_seconds": round(uptime),
        "request_count": monitor["request_count"],
        "last_request_time": monitor["last_request_time"],
        "total_messages_archived": monitor["total_messages_archived"],
        "last_message_time": monitor["last_message_time"],
        "error_count": len(monitor["errors"]),
        "recent_errors": monitor["errors"][-5:],
        "target_group_id": Config.TARGET_GROUP_ID,
    })


@app.route("/sync", methods=["POST"])
def trigger_sync():
    today = datetime.now(TAIPEI_TZ).strftime("%Y%m%d")
    path = Config.ARCHIVE_DIR / f"line_archive_{today}.xlsx"
    ok = push_xlsx(path, today, force=True)
    return jsonify({"status": "ok" if ok else "error"})


@app.route("/debug-imports", methods=["GET"])
def debug_imports():
    results = {}
    try:
        from archiver import get_sheet_rows, _get_sheet_name, SHEET_CRITICAL, SHEET_WARNING, SHEET_OTHERS
        results["archiver"] = "ok"
    except Exception as e:
        results["archiver"] = str(e)
    try:
        from sync_repo import push_xlsx, pull_xlsx
        results["sync_repo"] = "ok"
    except Exception as e:
        results["sync_repo"] = str(e)
    try:
        from llm_analyzer import analyze_messages
        results["llm_analyzer"] = "ok"
    except Exception as e:
        results["llm_analyzer"] = str(e)
    try:
        import openai
        results["openai"] = "ok"
    except Exception as e:
        results["openai"] = str(e)
    return jsonify(results)


@app.route("/debug-profile/<user_id>", methods=["GET"])
def debug_profile(user_id: str):
    from archiver import get_user_display_name
    name = get_user_display_name(user_id)
    return jsonify({"user_id": user_id, "display_name": name})


@app.route("/process", methods=["POST"])
def process_archive():
    try:
        today = datetime.now(TAIPEI_TZ).strftime("%Y%m%d")
        local_path = Config.ARCHIVE_DIR / f"line_archive_{today}.xlsx"

        pull_xlsx(today, local_path)

        critical = get_sheet_rows(local_path, SHEET_CRITICAL)
        warning = get_sheet_rows(local_path, SHEET_WARNING)
        others = get_sheet_rows(local_path, SHEET_OTHERS)

        from archiver import split_message, _get_sheet_name
        all_rows = {}
        for sheet, rows in [(SHEET_CRITICAL, critical), (SHEET_WARNING, warning), (SHEET_OTHERS, others)]:
            for r in rows:
                ts = r[0]
                name = r[1] if len(r) > 1 else ""
                uid = r[2] if len(r) > 2 else ""
                if sheet in (SHEET_CRITICAL, SHEET_WARNING):
                    if len(r) >= 5:
                        prefix, content = r[3], r[4]
                    elif len(r) >= 4:
                        prefix, content = split_message(r[3])
                    else:
                        continue
                    key = (ts, prefix, content)
                    all_rows[key] = [ts, name, uid, prefix, content, sheet]
                else:
                    raw_msg = r[3] if len(r) >= 4 else ""
                    real_sheet = _get_sheet_name(raw_msg)
                    if real_sheet in (SHEET_CRITICAL, SHEET_WARNING):
                        prefix, content = split_message(raw_msg)
                        key = (ts, prefix, content)
                        all_rows[key] = [ts, name, uid, prefix, content, real_sheet]
                    else:
                        key = (ts, raw_msg)
                        all_rows[key] = [ts, name, uid, raw_msg, sheet]

        sorted_rows = sorted(all_rows.values(), key=lambda r: r[0])

        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)
        bold = Font(bold=True)
        cols5 = ["timestamp", "sender_name", "sender_user_id", "prefix", "message"]
        cols4 = ["timestamp", "sender_name", "sender_user_id", "message"]
        sheets_data = {SHEET_CRITICAL: [], SHEET_WARNING: [], SHEET_OTHERS: []}
        for r in sorted_rows:
            sheet = r[-1]
            sheets_data[sheet].append(r)
        for sname in (SHEET_CRITICAL, SHEET_WARNING):
            ws = wb.create_sheet(title=sname)
            ws.append(cols5)
            for ci in range(1, 6):
                ws.cell(row=1, column=ci).font = bold
            for r in sheets_data[sname]:
                ws.append(r[:5])
        ws = wb.create_sheet(title=SHEET_OTHERS)
        ws.append(cols4)
        for ci in range(1, 5):
            ws.cell(row=1, column=ci).font = bold
        for r in sheets_data[SHEET_OTHERS]:
            ws.append(r[:4])

        wb.save(str(local_path))
        logger.info("Dedup saved: %d critical, %d warning, %d others (%d unique)",
                    len(sheets_data[SHEET_CRITICAL]), len(sheets_data[SHEET_WARNING]),
                    len(sheets_data[SHEET_OTHERS]), len(sorted_rows))

        push_xlsx(local_path, today, force=True)

        critical_texts = [f"{r[3]}，{r[4]}" for r in sheets_data[SHEET_CRITICAL]]
        warning_texts = [f"{r[3]}，{r[4]}" for r in sheets_data[SHEET_WARNING]]
        analysis = analyze_messages(critical_texts, warning_texts)

        return jsonify({
            "status": "ok",
            "date": today,
            "unique_messages": len(sorted_rows),
            "critical": len(sheets_data[SHEET_CRITICAL]),
            "warning": len(sheets_data[SHEET_WARNING]),
            "others": len(sheets_data[SHEET_OTHERS]),
            "analysis": analysis,
        })
    except Exception as e:
        logger.exception("/process failed")
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/send-report", methods=["POST"])
def trigger_report():
    today = datetime.now(TAIPEI_TZ).strftime("%Y%m%d")
    push_xlsx(Config.ARCHIVE_DIR / f"line_archive_{today}.xlsx", today, force=True)
    success = send_report(today)
    send_line_report(today)
    clear_archive(today)
    return jsonify({"status": "ok" if success else "email_failed_line_sent", "date": today})


@handler.add(JoinEvent)
def handle_join(event: JoinEvent) -> None:
    source = event.source
    group_id = getattr(source, "group_id", None)
    if group_id:
        logger.info("[GROUP ID DETECTED] groupId=%s (via join event)", group_id)
        if group_id not in detected_group_ids:
            detected_group_ids.append(group_id)

@handler.add(MessageEvent, message=TextMessageContent)
def handle_text_message(event: MessageEvent) -> None:
    message: TextMessageContent = event.message
    source = event.source
    group_id = getattr(source, "group_id", None)
    if group_id:
        logger.info("[GROUP ID DETECTED] groupId=%s", group_id)
        if group_id not in detected_group_ids:
            detected_group_ids.append(group_id)
        last_messages.append({
            "group_id": group_id,
            "user_id": source.user_id,
            "text": message.text[:100],
            "timestamp": datetime.fromtimestamp(event.timestamp / 1000.0, tz=TAIPEI_TZ).isoformat(),
        })
    if Config.TARGET_GROUP_ID and group_id != Config.TARGET_GROUP_ID:
        logger.debug("Ignoring message from non-target group: %s", group_id)
        return
    if not Config.TARGET_GROUP_ID and not group_id:
        logger.debug("Ignoring non-group message (no TARGET_GROUP_ID set)")
        return
    timestamp = datetime.fromtimestamp(event.timestamp / 1000.0, tz=TAIPEI_TZ).strftime("%Y-%m-%d %H:%M:%S")
    try:
        display_name = get_user_display_name(event.source.user_id)
        sheet = archive_message(
            timestamp=timestamp,
            sender_name=display_name,
            sender_user_id=event.source.user_id or "unknown",
            message=message.text,
        )
        monitor["total_messages_archived"] += 1
        monitor["last_message_time"] = datetime.now(TAIPEI_TZ).isoformat()
        logger.info("Archived message from %s (%s) to sheet [%s]", display_name, event.source.user_id, sheet)

        today_str = datetime.now(TAIPEI_TZ).strftime("%Y%m%d")
        push_xlsx(Config.ARCHIVE_DIR / f"line_archive_{today_str}.xlsx", today_str)

        # 檢查即時觸發條件 (critical + warning 合計)
        if sheet in (SHEET_CRITICAL, SHEET_WARNING):
            monitor["critical_count"] += 1
            if monitor["critical_count"] >= 3:
                trigger_instant_report(today_str, "critical+warning", monitor["critical_count"])
                monitor["critical_count"] = 0
                monitor["last_trigger_time"] = datetime.now(TAIPEI_TZ).isoformat()
    except Exception as e:
        monitor["errors"].append(str(e))
        logger.exception("Failed to archive message")


def init_scheduler() -> None:
    scheduler.add_job(
        func=scheduled_report_job,
        trigger="cron",
        hour=20,
        minute=0,
        timezone=TAIPEI_TZ,
        id="daily_report",
        replace_existing=True,
    )
    scheduler.start()
    logger.info("Scheduler started: daily report at 20:00 Asia/Taipei")


def create_app() -> Flask:
    Config.ensure_dirs()
    missing = Config.validate()
    if missing:
        logger.warning("Missing environment variables: %s", ", ".join(missing))
    init_scheduler()
    return app


if __name__ == "__main__":
    create_app()
    port = Config.APP_PORT
    logger.info("Starting local server on 0.0.0.0:%d", port)
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)




