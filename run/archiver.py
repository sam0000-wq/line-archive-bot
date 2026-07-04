# -*- coding: utf-8 -*-
"""Message classification and xlsx archiving module."""

import logging
import requests as http_requests
from datetime import datetime
from pathlib import Path
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
import re
import re
from config import Config

logger = logging.getLogger(__name__)
CRITICAL_COLUMNS = ["timestamp", "sender_name", "prefix", "message"]
OTHERS_COLUMNS = ["timestamp", "sender_name", "message"]
LLM_COLUMNS = ["timestamp", "analysis"]
SHEET_CRITICAL = "critical"
SHEET_WARNING = "warning"
SHEET_REPORT = "report"
SHEET_OTHERS = "others"
SHEET_LLM = "LLM"
SHEET_CHART = "chart"
PREFIX_SHEETS = (SHEET_CRITICAL, SHEET_WARNING, SHEET_REPORT)

_profile_cache: dict[str, str] = {}
DELIMITERS = (",", "，")


def _daily_path(date: Optional[datetime] = None) -> Path:
    if date is None:
        date = datetime.now()
    filename = f"line_archive_{date.strftime('%Y%m%d')}.xlsx"
    Config.ensure_dirs()
    return Config.ARCHIVE_DIR / filename


def _get_sheet_name(prefix: str) -> str:
    cleaned = prefix.lower().strip()
    if cleaned.startswith("critical"):
        return SHEET_CRITICAL
    elif cleaned.startswith("warning"):
        return SHEET_WARNING
    elif cleaned.startswith("report"):
        return SHEET_REPORT
    return SHEET_OTHERS


def get_user_display_name(user_id: str) -> str:
    if user_id in _profile_cache:
        return _profile_cache[user_id]
    if not Config.LINE_CHANNEL_ACCESS_TOKEN:
        logger.warning("No LINE_CHANNEL_ACCESS_TOKEN, using user_id")
        _profile_cache[user_id] = user_id
        return user_id
    try:
        url = f"https://api.line.me/v2/bot/profile/{user_id}"
        headers = {"Authorization": f"Bearer {Config.LINE_CHANNEL_ACCESS_TOKEN}"}
        resp = http_requests.get(url, headers=headers, timeout=5)
        logger.info("LINE profile API %s -> HTTP %d", user_id, resp.status_code)
        if resp.status_code == 200:
            display_name = resp.json().get("displayName", user_id)
            _profile_cache[user_id] = display_name
            logger.info("Display name: %s -> %s", user_id, display_name)
            return display_name
        else:
            logger.warning("LINE profile API returned %d: %s", resp.status_code, resp.text[:200])
    except Exception as e:
        logger.exception("Failed to fetch LINE profile for %s: %s", user_id, e)
    _profile_cache[user_id] = user_id
    return user_id


def split_message(message: str) -> tuple[str, str]:
    for d in DELIMITERS:
        if d in message:
            parts = message.split(d, 1)
            return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""
    return "", message


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(str(path))
        for sn in wb.sheetnames:
            if sn == SHEET_LLM:
                continue
            ws = wb[sn]
            if ws.max_row >= 1:
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                if headers == ["timestamp", "sender_name", "sender_user_id", "message"]:
                    ws.cell(row=1, column=1).value = "timestamp"
                    ws.cell(row=1, column=2).value = "sender_name"
                    ws.delete_cols(3)
                    if sn in (SHEET_CRITICAL, SHEET_WARNING):
                        ws.cell(row=1, column=3).value = "prefix"
                        ws.cell(row=1, column=4).value = "message"
                        for row_idx in range(2, ws.max_row + 1):
                            old_msg = ws.cell(row=row_idx, column=3).value or ""
                            prefix, content = split_message(str(old_msg))
                            ws.cell(row=row_idx, column=3).value = prefix
                            ws.cell(row=row_idx, column=4).value = content
                    logger.info("Migrated sheet [%s] (removed sender_user_id)", sn)
        if SHEET_LLM not in wb.sheetnames:
            ws = wb.create_sheet(title=SHEET_LLM)
            ws.append(LLM_COLUMNS)
            bold = Font(bold=True)
            for ci in range(1, len(LLM_COLUMNS) + 1):
                ws.cell(row=1, column=ci).font = bold
        wb.save(str(path))
        return load_workbook(str(path))

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    bold = Font(bold=True)
    for sheet_name, cols in [(SHEET_CRITICAL, CRITICAL_COLUMNS), (SHEET_WARNING, CRITICAL_COLUMNS), (SHEET_REPORT, CRITICAL_COLUMNS), (SHEET_OTHERS, OTHERS_COLUMNS)]:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(cols)
        for col_idx in range(1, len(cols) + 1):
            ws.cell(row=1, column=col_idx).font = bold
    ws_llm = wb.create_sheet(title=SHEET_LLM)
    ws_llm.append(LLM_COLUMNS)
    for col_idx in range(1, len(LLM_COLUMNS) + 1):
        ws_llm.cell(row=1, column=col_idx).font = bold
    wb.save(str(path))
    return load_workbook(str(path))


def get_sheet_rows(path: Path, sheet_name: str) -> list[list[str]]:
    if not path.exists():
        return []
    try:
        wb = load_workbook(str(path))
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            rows.append([str(c) if c is not None else "" for c in row])
        return rows
    except Exception:
        logger.exception("Failed to read sheet %s from %s", sheet_name, path)
        return []


def archive_message(timestamp: str, sender_name: str, message: str) -> str:
    sheet_name = _get_sheet_name(message)
    path = _daily_path()
    try:
        wb = _ensure_workbook(path)
        ws = wb[sheet_name]
        if sheet_name in PREFIX_SHEETS:
            prefix, content = split_message(message)
            ws.append([timestamp, sender_name, prefix, content])
        else:
            ws.append([timestamp, sender_name, message])
        wb.save(str(path))
        logger.info("Archived to [%s] %s", sheet_name, path.name)
        return sheet_name
    except Exception:
        logger.exception("Failed to archive message to %s", path)
        raise


def write_llm_analysis(path: Path, analysis: str) -> None:
    try:
        wb = load_workbook(str(path))
        if SHEET_LLM not in wb.sheetnames:
            ws = wb.create_sheet(title=SHEET_LLM)
            ws.append(LLM_COLUMNS)
            bold = Font(bold=True)
            for ci in range(1, len(LLM_COLUMNS) + 1):
                ws.cell(row=1, column=ci).font = bold
        else:
            ws = wb[SHEET_LLM]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, analysis])
        wb.save(str(path))
        logger.info("LLM analysis written to %s", path.name)
    except Exception:
        logger.exception("Failed to write LLM analysis to %s", path)


def clear_archive(date_str: Optional[str] = None) -> bool:
    path = _daily_path() if not date_str else Config.ARCHIVE_DIR / f"line_archive_{date_str}.xlsx"
    if not path.exists():
        return True
    try:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)
        bold = Font(bold=True)
        for sheet_name, cols in [(SHEET_CRITICAL, CRITICAL_COLUMNS), (SHEET_WARNING, CRITICAL_COLUMNS), (SHEET_REPORT, CRITICAL_COLUMNS), (SHEET_OTHERS, OTHERS_COLUMNS)]:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(cols)
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=1, column=col_idx).font = bold
        ws_llm = wb.create_sheet(title=SHEET_LLM)
        ws_llm.append(LLM_COLUMNS)
        for col_idx in range(1, len(LLM_COLUMNS) + 1):
            ws_llm.cell(row=1, column=col_idx).font = bold
        wb.save(str(path))
        logger.info("Cleared archive: %s", path.name)
        return True
    except Exception:
        logger.exception("Failed to clear archive %s", path)
        return False


def get_archive_path(date_str: Optional[str] = None) -> Optional[Path]:
    if date_str:
        path = Config.ARCHIVE_DIR / f"line_archive_{date_str}.xlsx"
    else:
        path = _daily_path()
    return path if path.exists() else None


STATION_KEYWORDS = {
    "組裝": "組裝站",
    "SMT": "組裝站",
    "DIP": "組裝站",
    "焊接": "組裝站",
    "測試": "測試站",
    "Burn-in": "測試站",
    "stress": "測試站",
    "BIOS": "測試站",
    "DIMM": "測試站",
    "GPU": "測試站",
    "包裝": "包裝站",
    "出貨": "包裝站",
    "品檢": "包裝站",
    "外觀": "包裝站",
    "物料": "物料站",
    "進貨": "物料站",
    "來料": "物料站",
    "庫存": "物料站",
    "倉庫": "物料站",
    "發料": "物料站",
}


def _extract_station(message: str) -> str:
    for keyword, station in STATION_KEYWORDS.items():
        if keyword.lower() in message.lower():
            return station
    return "其他"


def _parse_llm_analysis(llm_rows: list[list[str]]) -> dict:
    """Parse LLM analysis text to extract priority and unit counts."""
    priority_counts = {"P1": 0, "P2": 0, "P3": 0}
    unit_counts: dict[str, int] = {}
    units_pattern = re.compile(r"(組裝站|測試站|包裝站|物料站|組裝|測試|包裝|物料)")

    for row in llm_rows:
        text = row[1] if len(row) >= 2 else ""

        p1 = len(re.findall(r"P1", text))
        p2 = len(re.findall(r"P2", text))
        p3 = len(re.findall(r"P3", text))
        priority_counts["P1"] += p1
        priority_counts["P2"] += p2
        priority_counts["P3"] += p3

        action_section = False
        for line in text.split("\n"):
            if "行動方案" in line:
                action_section = True
            elif re.match(r"【.+】", line):
                action_section = False
            if action_section:
                matches = units_pattern.findall(line)
                for m in matches:
                    unit = m if m.endswith("站") else m + "站"
                    unit_counts[unit] = unit_counts.get(unit, 0) + 1

    if sum(priority_counts.values()) == 0:
        for line in (row[1] for row in llm_rows if len(row) >= 2):
            priority_counts["P1"] += len(re.findall(r"P1", line))
            priority_counts["P2"] += len(re.findall(r"P2", line))
            priority_counts["P3"] += len(re.findall(r"P3", line))

    return {"priority": priority_counts, "units": unit_counts}


def create_chart_sheet(path: Path) -> None:
    try:
        wb = load_workbook(str(path))

        critical = get_sheet_rows(path, SHEET_CRITICAL)
        warning = get_sheet_rows(path, SHEET_WARNING)
        report = get_sheet_rows(path, SHEET_REPORT)
        others = get_sheet_rows(path, SHEET_OTHERS)
        llm_rows = get_sheet_rows(path, SHEET_LLM) if SHEET_LLM in wb.sheetnames else []

        llm_data = _parse_llm_analysis(llm_rows) if llm_rows else {"priority": {"P1": 0, "P2": 0, "P3": 0}, "units": {}}

        type_counts = {
            "CRITICAL": len(critical),
            "WARNING": len(warning),
            "REPORT": len(report),
            "OTHERS": len(others),
        }

        station_counts: dict[str, int] = {}
        for sheet_rows in [critical, warning, report, others]:
            for r in sheet_rows:
                msg = r[3] if len(r) >= 4 else (r[2] if len(r) >= 3 else "")
                station = _extract_station(msg)
                station_counts[station] = station_counts.get(station, 0) + 1

        if SHEET_CHART in wb.sheetnames:
            del wb[SHEET_CHART]
        ws = wb.create_sheet(title=SHEET_CHART)

        bold = Font(bold=True, size=12)
        ws["A1"] = "訊息分類統計"
        ws["A1"].font = bold
        ws["A2"] = "類別"
        ws["B2"] = "數量"
        ws["A2"].font = Font(bold=True)
        ws["B2"].font = Font(bold=True)
        row = 3
        for label, count in type_counts.items():
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            row += 1

        ws["D1"] = "站點統計"
        ws["D1"].font = bold
        ws["D2"] = "站點"
        ws["E2"] = "數量"
        ws["D2"].font = Font(bold=True)
        ws["E2"].font = Font(bold=True)
        row = 3
        for station, count in sorted(station_counts.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=4, value=station)
            ws.cell(row=row, column=5, value=count)
            row += 1

        ws["G1"] = "LLM 分析：優先順序分布"
        ws["G1"].font = bold
        ws["G2"] = "優先順序"
        ws["H2"] = "數量"
        ws["G2"].font = Font(bold=True)
        ws["H2"].font = Font(bold=True)
        row = 3
        for p_label, p_count in llm_data["priority"].items():
            ws.cell(row=row, column=7, value=p_label)
            ws.cell(row=row, column=8, value=p_count)
            row += 1

        ws["J1"] = "LLM 分析：負責單位分布"
        ws["J1"].font = bold
        ws["J2"] = "單位"
        ws["K2"] = "數量"
        ws["J2"].font = Font(bold=True)
        ws["K2"].font = Font(bold=True)
        row = 3
        for unit, count in sorted(llm_data["units"].items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=10, value=unit)
            ws.cell(row=row, column=11, value=count)
            row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 10

        pie = PieChart()
        pie.title = "訊息分類占比"
        pie.style = 10
        data = Reference(ws, min_col=2, min_row=2, max_row=6)
        cats = Reference(ws, min_col=1, min_row=3, max_row=6)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.width = 16
        pie.height = 10
        ws.add_chart(pie, "A8")

        bar = BarChart()
        bar.title = "各站點訊息數"
        bar.style = 10
        bar.type = "col"
        bar.y_axis.title = "數量"
        station_row_end = 2 + len(station_counts)
        data2 = Reference(ws, min_col=5, min_row=2, max_row=station_row_end)
        cats2 = Reference(ws, min_col=4, min_row=3, max_row=station_row_end)
        bar.add_data(data2, titles_from_data=True)
        bar.set_categories(cats2)
        bar.width = 16
        bar.height = 10
        ws.add_chart(bar, "J8")

        if sum(llm_data["priority"].values()) > 0:
            pie2 = PieChart()
            pie2.title = "優先順序分布 (P1/P2/P3)"
            pie2.style = 10
            data3 = Reference(ws, min_col=8, min_row=2, max_row=5)
            cats3 = Reference(ws, min_col=7, min_row=3, max_row=5)
            pie2.add_data(data3, titles_from_data=True)
            pie2.set_categories(cats3)
            pie2.width = 16
            pie2.height = 10
            ws.add_chart(pie2, "A24")

        if llm_data["units"]:
            bar2 = BarChart()
            bar2.title = "負責單位分布 (LLM 分析)"
            bar2.style = 10
            bar2.type = "col"
            bar2.y_axis.title = "行動項數"
            unit_row_end = 2 + len(llm_data["units"])
            data4 = Reference(ws, min_col=11, min_row=2, max_row=unit_row_end)
            cats4 = Reference(ws, min_col=10, min_row=3, max_row=unit_row_end)
            bar2.add_data(data4, titles_from_data=True)
            bar2.set_categories(cats4)
            bar2.width = 16
            bar2.height = 10
            ws.add_chart(bar2, "J24")

        wb.save(str(path))
        logger.info("Chart sheet created in %s", path.name)
    except Exception:
        logger.exception("Failed to create chart sheet in %s", path)
